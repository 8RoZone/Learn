from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackQueryHandler
from telegram import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from openpyxl import Workbook, load_workbook
import datetime
import pyperclip

pathExcel = "../data/data_telegram.xlsx"
BTN_IN_CHAT, BTN_SHARE, BTN_PLAY_CHANNEL = range(3)
# Hàm xử lý khi người dùng bấm start
def start(update, context):
    # tải workbook và gán các worksheet
    wb = load_workbook(pathExcel)
    sheet_giftcode = wb["Giftcode"]
    sheet_data_username = wb["DataUsername"]
    sheet_data_phone = wb["DataPhone"]
    sheet_data_new = wb["DataNew"]
    sheet_data_crazy = wb["DataCrazy"]

    user = update.effective_user
    username = user.username
    phone_number = "null"
    
    # check danh sách đã phát, nếu có tạo các btn điều hướng
    if check_generate(username,sheet_data_crazy)[0]:
        reply_markup = create_button(BTN_IN_CHAT)
        update.message.reply_text('Nhấn các nút điều hướng:', reply_markup=reply_markup)
        return
    # check danh sách chờ xét duyệt khi không có trong danh sách
    if check_generate(username,sheet_data_new)[0]:
        reply_markup = create_button(BTN_IN_CHAT)
        update.message.reply_text('Xin chờ xét duyệt!\nNhấn các nút điều hướng:', reply_markup=reply_markup)
        return

    # Kiểm tra xem username có trong danh sách hay không
    if check_generate(username,sheet_data_username)[0]:
        cell_position = check_generate(username,sheet_data_username)[1]
        column_index = cell_position.column  #vị trí cột chứa username trùng
        row_index = cell_position.row        #vị trí hàng chứa username trùng
        if sheet_data_username[f'B{row_index}'].value == None: # kiểm tra đã phát chưa (nếu có sẽ lưu code tại col B)
            # truyền vào wb để lưu trong hàm wb.save
            # sheet_giftcode lấy giftcode
            # sheet_data_username lưu code đã phát tại B['row_index']
            giftcode = get_giftcode(wb, sheet_giftcode, sheet_data_username, row_index)
            # kiểm tra code còn không
            if giftcode == False:
                message = "Đã phát hết, mai đến sớm nhé!!"
                context.bot.send_message(chat_id=user.id, text=message)
                return
            # thông báo code cho người dùng
            message = "Bạn có trong ds.\n"
            message += f"Giftcode nè: {giftcode}"
            # lưu lại thông tin để so sánh ban đầu
            data_crazy = [username, phone_number, giftcode, datetime.datetime.now()]
            last_row = sheet_data_crazy.max_row + 1
            for i, value in enumerate(data_crazy, start=1):
                sheet_data_crazy.cell(row=last_row, column=i, value=value)
            wb.save(pathExcel)

            reply_markup = create_button(BTN_IN_CHAT)
            update.message.reply_text(f'{message} \nNhấn các nút điều hướng:', reply_markup=reply_markup)
            reply_markup = create_button(BTN_PLAY_CHANNEL)
            update.message.reply_text('Chúc bạn chơi game vui vẻ!!', reply_markup=reply_markup)
            return
        else:
            reply_markup = create_button(BTN_IN_CHAT)
            update.message.reply_text(f'Mỗi người chỉ nhận được 1 code!', reply_markup=reply_markup)
            return
    else:
        reply_markup = create_button(BTN_SHARE)
        message = f"Bạn không có trong danh sách.\n Hãy chia sẻ số điện thoại để kiểm tra."
        context.bot.send_message(chat_id=user.id, text=message, reply_markup=reply_markup)
        return
    


def check_generate(valueCompare, sheetCompare):
    result = False
    position = 0
    # lưu username/phone_number ở cột A
    for cell in sheetCompare['A']:
        str1 = str(valueCompare)
        str2 = str(cell.value)
        if str1 == str2:
            position = cell
            result = True
    return [result,position]

def get_giftcode(wb, sheet_giftcode, sheetSave, row_index):
    last_row = sheet_giftcode.max_row
    giftcode = sheet_giftcode[f'A{last_row}'].value
    if last_row == 1:
        return False
    print(giftcode)
    print(sheetSave)
    sheetSave[f'B{row_index}'] = giftcode
    print(giftcode)
    print(sheetSave[f'B{row_index}'].value)

    sheetSave[f'C{row_index}'] = datetime.datetime.now()
    sheet_giftcode[f'A{last_row}'] = None
    wb.save(pathExcel)
    return giftcode

def share_contact(update, context):
    user = update.effective_user
    contact = update.effective_message.contact
    # Lấy số điện thoại từ contact
    phone_number = contact.phone_number
    
    wb = load_workbook(pathExcel)
    sheet_giftcode = wb["Giftcode"]
    sheet_data_username = wb["DataUsername"]
    sheet_data_phone = wb["DataPhone"]
    sheet_data_new = wb["DataNew"]
    sheet_data_crazy = wb["DataCrazy"]

    # Kiểm tra xem username có trong danh sách hay không
    if check_generate(phone_number,sheet_data_phone)[0]:
        cell_position = check_generate(phone_number,sheet_data_phone)[1]
        column_index = cell_position.column
        row_index = cell_position.row
        if sheet_data_phone[f'B{row_index}'].value == None:
            giftcode = get_giftcode(wb, sheet_giftcode, sheet_data_phone, row_index)
            if giftcode == False:
                message = "Đã phát hết, mai đến sớm nhé!!"
                context.bot.send_message(chat_id=user.id, text=message)
                return
            message = "Bạn có trong ds.\n"
            message += f"Giftcode nè: {giftcode}"
            data_crazy = [user.username, phone_number, giftcode, datetime.datetime.now()]
            last_row = sheet_data_crazy.max_row + 1
            for i, value in enumerate(data_crazy, start=1):
                sheet_data_crazy.cell(row=last_row, column=i, value=value)
            wb.save(pathExcel)
            reply_markup = create_button(BTN_IN_CHAT)
            update.message.reply_text(f'{message} \nNhấn các nút điều hướng:', reply_markup=reply_markup)
            reply_markup = create_button(BTN_PLAY_CHANNEL)
            update.message.reply_text('Chúc bạn chơi game vui vẻ!!', reply_markup=reply_markup)
            return
        else:
            message = f'Mỗi người chỉ nhận 1 code.\nPlay để chơi, Channel để xem thông tin chơi!'
            reply_markup = create_button(BTN_PLAY_CHANNEL)
            context.bot.send_message(chat_id=user.id, text=message, reply_markup=reply_markup)
            return False
    else:
        last_row = sheet_data_new.max_row + 1
        dataUser = [user.username, phone_number, datetime.datetime.now()]
        if check_generate(user.username, sheet_data_new)[0] == False:
            for i, value in enumerate(dataUser, start=1):
                sheet_data_new.cell(row=last_row, column=i, value=value)
        message = "Chờ duyệt nhé em."
        wb.save(pathExcel)
        reply_markup = create_button(BTN_IN_CHAT)
        update.message.reply_text('Xin chờ xét duyệt!\nNhấn các nút điều hướng:', reply_markup=reply_markup)
        reply_markup = create_button(BTN_PLAY_CHANNEL)
        update.message.reply_text('Trong thời gian chờ, bạn có thể nhấn Play để trải nghiệm!!', reply_markup=reply_markup)
        return


def create_button(type):
    if type == BTN_IN_CHAT:
        button1 = InlineKeyboardButton("YouTube", url="https://www.youtube.com")
        button2 = InlineKeyboardButton("Google", url="https://www.google.com")
        button3 = InlineKeyboardButton("Button 3", callback_data="button3")
        button4 = InlineKeyboardButton("Button 4", callback_data="button4")
        button5 = InlineKeyboardButton("Button 5", callback_data="button5")
        reply_markup = InlineKeyboardMarkup([
            [button1, button2],
            [button3, button4],
            [button5]
        ])
        return reply_markup
    
    if type == BTN_SHARE:
        button = KeyboardButton(text="Chia sẻ số điện thoại", request_contact=True)
        reply_markup = ReplyKeyboardMarkup([[button]], resize_keyboard=True, one_time_keyboard=True)
        return reply_markup

    if type == BTN_PLAY_CHANNEL:
        button1 = KeyboardButton(text="Play")
        button2 = KeyboardButton(text="Channel")
        reply_markup = ReplyKeyboardMarkup([[button1, button2]], resize_keyboard=True, one_time_keyboard=True)
        return reply_markup



def handle_button_click(update, context):
    # Xử lý tin nhắn từ người dùng
    # Xử lý nút được nhấn
    query = update.callback_query
    query.answer()
    button_data = query.data
    query.message.reply_text(f'Bạn đã nhấn nút "{button_data}".')

# Hàm xử lý sự kiện khi người dùng gửi tin nhắn
def handle_message(update, context):
    # Xử lý tin nhắn từ người dùng
    message = update.message.text
    # Kiểm tra nút được nhấn
    if message == 'Play':
        update.message.reply_text('Bạn đã chọn Button 1')
    elif message == 'Channel':
        update.message.reply_text('Bạn đã chọn Button 2')

# Hàm main
def main():
    # Khởi tạo API Token của bot Telegram
    TOKEN = '6343669796:AAFcB4Vq4PmCzr4W1ZJQwAYyjMWLwwwIW-k'
    # Khởi tạo đối tượng updater
    updater = Updater(token=TOKEN, use_context=True)
    
    # Lấy đối tượng dispatcher
    dispatcher = updater.dispatcher

    # Đăng ký command handler cho start
    start_handler = CommandHandler('start', start)
    dispatcher.add_handler(start_handler)
    
    button_handler = CallbackQueryHandler(handle_button_click)
    dispatcher.add_handler(button_handler)

   

    share_contact_handler = MessageHandler(Filters.contact, share_contact)
    dispatcher.add_handler(share_contact_handler)


    message_handler = MessageHandler(Filters.text, handle_message)
    dispatcher.add_handler(message_handler)


    # Bắt đầu lắng nghe các tin nhắn mới
    updater.start_polling()
    
    # Dừng chương trình khi nhận được tín hiệu Ctrl-C
    updater.idle()

# Gọi hàm main để chạy bot
if __name__ == '__main__':
    main()