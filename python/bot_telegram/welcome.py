from telegram import Update
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackContext
from telegram.ext import ConversationHandler, CallbackQueryHandler
from telegram import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import datetime

wb = load_workbook('../data/data_telegram.xlsx')
ws = wb.active
# Tìm hàng cuối cùng trong cột A (cột 1)
last_row = ws.max_row + 1
dataUser = []

STATE_1, STATE_2, STATE_3, STATE_4, STATE_5 = range(5)

def start(update: Update, context: CallbackContext):
    user = update.effective_user
    if user:
        id = user.id
        first_name = user.first_name
        username = user.username
        dataUser.append(id)
        dataUser.append(first_name)
        dataUser.append(username)
        for i, value in enumerate(dataUser, start=1):
            ws.cell(row=last_row, column=i, value=value)
        wb.save("../data/data_telegram.xlsx")
        print(user)
        message = f"\U0001F3C6 Xin chào {first_name}!\U0001F3C6\n"
        message += "Your info:\n"
        message += f"\u2705 id: {id}\n"
        message += f"\u2705 username: {username}\n"
        context.bot.send_message(chat_id=update.effective_chat.id, text=message)
    else:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Xin chào!")
    MessageHandler(Filters.text, share_info)
    return STATE_1
def cancel(update: Update, context: CallbackContext):
    update.message.reply_text('See you again!')
    return ConversationHandler.END


def share_info(update: Update, context: CallbackContext):
    print("share_info")
    keyboard = [
        [KeyboardButton("\u260E Chia sẻ thông tin", request_contact=True)]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
    update.message.reply_text('Vui lòng chia sẻ thông tin của bạn.', reply_markup=reply_markup)
    return STATE_2

def handle_info(update: Update, context: CallbackContext):
    phone_number = update.message.contact.phone_number
    ws.cell(row=last_row, column=4, value=datetime.datetime.now())
    ws.cell(row=last_row, column=5, value=phone_number)
    wb.save("../data/data_telegram.xlsx")
    update.message.reply_text(f"Số điện thoại của bạn là: {phone_number}")
    return STATE_3
def chat_3(update, context):
    # Xử lý bước chat 3
    update.message.reply_text('chat_3!')
    return STATE_4
def chat_4(update, context):
    # Xử lý bước chat 4
    update.message.reply_text('chat_4!')
    return STATE_5


def chat_5(update, context):
    # Xử lý bước chat 5
    update.message.reply_text('chat_5!')
    
def main():
    updater = Updater(token='6343669796:AAFcB4Vq4PmCzr4W1ZJQwAYyjMWLwwwIW-k', use_context=True)
    dispatcher = updater.dispatcher

    conv_handler = ConversationHandler(
    entry_points=[CommandHandler('start', start)],
    states={
        STATE_1: [MessageHandler(Filters.text, share_info)],
        STATE_2: [MessageHandler(Filters.contact, handle_info)],
        STATE_3: [MessageHandler(Filters.text, chat_3)],
        STATE_4: [MessageHandler(Filters.text, chat_4)],
        STATE_5: [MessageHandler(Filters.text, chat_5)],
    },
    fallbacks=[CommandHandler('cancel', cancel)],

    )
    
    
    dispatcher.add_handler(conv_handler)
    # dispatcher.add_handler(CommandHandler("start", start))

    updater.start_polling()
    updater.idle()

if __name__ == '__main__':
    main()


# user
# {'id': 6185977946, 'first_name': 'Ro', 'is_bot': False, 'username': 'oro3568', 'language_code': 'en'}
