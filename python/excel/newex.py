from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# name sheet
ws.title = "New Title" 
# create new sheets
ws1 = wb.create_sheet("Mysheet1", 1)
ws2 = wb.create_sheet("Mysheet2", 0)
ws3 = wb.create_sheet("Mysheet3", -1)


# Data can be assigned directly to cells
ws1['A2'] = 45 

# Rows can also be appended
ws.append([1, 2, 3])
ws.append([1, 2, 3])


# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

source = wb.active
target = wb.copy_worksheet(source)
print(target)

ws['A4'] = 4
c = ws['A4']
print(c)
d = ws.cell(row=4, column=2, value=10)

for x in range(1,101):
    for y in range(1,101):
        ws.cell(row=x, column=y)


# pham vi cell_range
cell_range = ws['A5':'C7']
for row in cell_range: 
    for cell in row:
        cell.value = 1
        print(cell.value)


ws2['C9'] = 'hello world'
a = (tuple(ws2.rows))
for x in a:
    print(x)
    for y in x:
        print(y.value)

for row in ws2.values:
   for value in row:
     print(value)



my_list = [1, 2, 3, 4, 5]
my_iter = iter(my_list)
print(next(my_iter))  # In ra 1
print(next(my_iter))  # In ra 2
print(next(my_iter))  # In ra 3
print(next(my_iter))  # In ra 4
print(next(my_iter))  # In ra 5
# print(next(my_iter))  # In ra StopIteration

# Save the file
wb.save("./data/sample.xlsx")