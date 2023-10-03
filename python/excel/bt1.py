from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active
ws.title = "BT1"

treeData = [["Type", "Leaf Color", "Height"], ["Maple", "Red", 549], ["Oak", "Green", 783], ["Pine", "Green", 1204]]
for x in treeData:
    ws.append(x)

cell_range = ws['A1':'C3']
for row in cell_range: 
    for cell in row:
        print(cell.value)

font = Font(name='Arial', size=12, bold=True, italic=False, underline='doubleAccounting',color='FF0000')

for row in ws["A1:C1"]:
    for cell in row:
        cell.font = font

wb.save("../data/bt1.xlsx")
